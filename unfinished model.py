# %%
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np


# %%
import pandas as pd
f_data = pd.read_excel(r"C:\Users\royr\Downloads\Sample_fraud_analytics_intern_v1.1.xlsx")

# %%
f_data.dtypes

# %%
f_data.shape

# %%
f_data.head()

# %%
f_data.nunique()

# %%
# --- Missing values check ---
missing_values = f_data.isnull().sum()

print("Missing values per column:\n", missing_values)

# %%
# f_data['total_price'] = f_data['total_price'].fillna(f_data['quantity'] * f_data['unit_price'])
# f_data['unit_price'] = f_data['unit_price'].fillna(f_data['total_price'] / f_data['quantity'])
# f_data['quantity'] = f_data['quantity'].fillna(f_data['total_price'] / f_data['unit_price'])
# f_data = f_data.dropna(subset=['quantity','unit_price','total_price'], how='all')
# # Drop rows where at least 2 of the 3 key numeric fields are missing
# f_data = f_data.dropna(
#     subset=['quantity','unit_price','total_price'],
#     thresh=2  # requires at least 2 non-missing values to keep the row
# )

# print("Remaining rows:", f_data.shape[0])


# %%
missing_values = f_data.isnull().sum()

print("Missing values per column:\n", missing_values)

# %%
f_data['quantity'] = pd.to_numeric(f_data['quantity'], errors='coerce')
f_data['unit_price'] = pd.to_numeric(f_data['unit_price'], errors='coerce')
f_data['total_price'] = pd.to_numeric(f_data['total_price'], errors='coerce')

# Record rows with missing amount fields before any later imputation or fill steps.
f_data['has_missing_amount_fields'] = f_data[['quantity', 'unit_price', 'total_price']].isna().any(axis=1)

# Impute algebraically only when exactly one of the three amount fields is missing.
amount_cols = ['quantity', 'unit_price', 'total_price']
missing_count = f_data[amount_cols].isna().sum(axis=1)

mask_quantity_missing = (
    (missing_count == 1)
    & f_data['quantity'].isna()
    & f_data['unit_price'].notna()
    & f_data['total_price'].notna()
)
mask_unit_price_missing = (
    (missing_count == 1)
    & f_data['unit_price'].isna()
    & f_data['quantity'].notna()
    & f_data['total_price'].notna()
)
mask_total_price_missing = (
    (missing_count == 1)
    & f_data['total_price'].isna()
    & f_data['quantity'].notna()
    & f_data['unit_price'].notna()
)

f_data.loc[mask_quantity_missing, 'quantity'] = (
    f_data.loc[mask_quantity_missing, 'total_price'] / f_data.loc[mask_quantity_missing, 'unit_price']
)
f_data.loc[mask_unit_price_missing, 'unit_price'] = (
    f_data.loc[mask_unit_price_missing, 'total_price'] / f_data.loc[mask_unit_price_missing, 'quantity']
)
f_data.loc[mask_total_price_missing, 'total_price'] = (
    f_data.loc[mask_total_price_missing, 'quantity'] * f_data.loc[mask_total_price_missing, 'unit_price']
)

print("Rows with missing amount fields:", f_data['has_missing_amount_fields'].sum())
print("Rows with exactly one amount field missing:", int((missing_count == 1).sum()))

# %%
f_data.describe(include=[np.number])

# %%
vendor_counts = f_data['vendor_id'].value_counts(dropna=False)
region_counts = f_data['region'].value_counts(dropna=False)
category_counts = f_data['category'].value_counts(dropna=False)

print("Vendor frequency counts:")
print(vendor_counts)

print("\nRegion frequency counts:")
print(region_counts)

print("\nCategory frequency counts:")
print(category_counts)
region_counts = f_data['region'].value_counts(dropna=False)
category_counts = f_data['category'].value_counts(dropna=False)

print("Vendor frequency counts:")
print(vendor_counts)

print("\nRegion frequency counts:")
print(region_counts)

print("\nCategory frequency counts:")
print(category_counts)

# %%
# vendor_counts = f_data['vendor_id'].value_counts()
# f_data['vendor_id_encoded'] = f_data['vendor_id'].map(vendor_counts)

# f_data = pd.get_dummies(f_data, columns=['category'], drop_first=True)

# f_data = pd.get_dummies(f_data, columns=['region'], drop_first=True)

# %%
# Boxplot for Quantity
plt.figure(figsize=(6,4))
plt.boxplot(f_data['quantity'].dropna(), vert=False, patch_artist=True,
            boxprops=dict(facecolor='lightblue', color='black'),
            medianprops=dict(color='red'))
plt.title("Quantity Outliers")
plt.xlabel("Quantity")
plt.show()


# Boxplot for Unit Price
plt.figure(figsize=(6,4))
plt.boxplot(f_data['unit_price'].dropna(), vert=False, patch_artist=True,
            boxprops=dict(facecolor='lightgreen', color='black'),
            medianprops=dict(color='red'))
plt.title("Unit Price Outliers")
plt.xlabel("Unit Price")
plt.show()


# Boxplot for Total Price
plt.figure(figsize=(6,4))
plt.boxplot(f_data['total_price'].dropna(), vert=False, patch_artist=True,
            boxprops=dict(facecolor='lightyellow', color='black'),
            medianprops=dict(color='red'))
plt.title("Total Price Outliers")
plt.xlabel("Total Price")
plt.show()


# %%

f_data['total'] = f_data['quantity'] * f_data['unit_price']


# %%
f_data.groupby('year')['total'].sum().plot(kind='bar', figsize=(8,5))
plt.title("Total Revenue by Year")
plt.xlabel("Year")
plt.ylabel("Revenue")
plt.show()

f_data.groupby('month')['total'].sum().plot(kind='bar', figsize=(8,5))
plt.title("Total Revenue by Month")
plt.xlabel("Month")
plt.ylabel("Revenue")
plt.show()


# %%

# Select only numeric columns
numeric_cols = f_data.select_dtypes(include=['float64','int64']).columns

# Compute stats
stats = pd.DataFrame({
    'std_dev': f_data[numeric_cols].std(),
    'skewness': f_data[numeric_cols].skew(),
    'kurtosis': f_data[numeric_cols].kurt()
})

print(stats)


# %%
# Aggregate by date
daily_counts = f_data.groupby('date').size()
daily_revenue = f_data.groupby('date')['total'].sum()

fig, ax1 = plt.subplots(figsize=(12,6))

# Plot transaction counts
ax1.plot(daily_counts.index, daily_counts.values, color='blue', marker='o', label='Transactions')
ax1.set_xlabel("Date")
ax1.set_ylabel("Number of Transactions", color='blue')
ax1.tick_params(axis='y', labelcolor='blue')

# Plot revenue on secondary axis
ax2 = ax1.twinx()
ax2.plot(daily_revenue.index, daily_revenue.values, color='green', marker='x', label='Revenue')
ax2.set_ylabel("Total Revenue", color='green')
ax2.tick_params(axis='y', labelcolor='green')

plt.title("Transactions and Revenue Over Time")
plt.grid(True)
plt.show()


# %%
import seaborn as sns
import pandas as pd

# Create pivot table: counts by day_of_week and hour
heatmap_data = f_data.groupby(['day_of_week','hour']).size().unstack(fill_value=0)

plt.figure(figsize=(10,6))
sns.heatmap(heatmap_data, cmap="YlGnBu", annot=False, cbar=True)
plt.title("Transaction Density by Day of Week and Hour")
plt.xlabel("Hour of Day (0-23)")
plt.ylabel("Day of Week (0=Monday)")
plt.show()


# %%

# --- Step 1: Make sure total exists ---
if 'total' not in f_data.columns:
    f_data['total'] = f_data['quantity'] * f_data['unit_price']

# --- Step 2: Select price-related columns ---
price_cols = ['unit_price', 'quantity', 'total']
price_data = f_data[price_cols]

# --- Step 3: Correlation matrix ---
corr_matrix = price_data.corr()

plt.figure(figsize=(6,4))
sns.heatmap(corr_matrix, annot=True, cmap="coolwarm", center=0)
plt.title("Correlation Heatmap (Price-related Columns)")
plt.show()

# --- Step 4: Scatter plots for pairwise relationships ---
sns.pairplot(price_data, diag_kind='hist')
plt.suptitle("Scatter Plots of Price-related Columns", y=1.02)
plt.show()



# %%

# Hour cyclical encoding (0–23)
f_data['hour_sin'] = np.sin(2 * np.pi * f_data['hour'] / 24)
f_data['hour_cos'] = np.cos(2 * np.pi * f_data['hour'] / 24)

# Day of week cyclical encoding (0–6)
f_data['dow_sin'] = np.sin(2 * np.pi * f_data['day_of_week'] / 7)
f_data['dow_cos'] = np.cos(2 * np.pi * f_data['day_of_week'] / 7)

# Check results
print(f_data[['hour','hour_sin','hour_cos','day_of_week','dow_sin','dow_cos']].head())


# %%
from sklearn.preprocessing import StandardScaler
scaler = StandardScaler()
X_scaled = scaler.fit_transform(f_data[['quantity','unit_price','total_price','hour_sin','hour_cos','dow_sin','dow_cos']])


# %%
# # One-hot encode region
# region_dummies = pd.get_dummies(f_data['region'], prefix='region', drop_first=False)
# f_data = pd.concat([f_data, region_dummies], axis=1)

# # One-hot encode category
# category_dummies = pd.get_dummies(f_data['category'], prefix='category', drop_first=False)
# f_data = pd.concat([f_data, category_dummies], axis=1)

# # Frequency encode vendor_id
# vendor_freq = f_data['vendor_id'].map(vendor_counts)
# f_data['vendor_id_freq'] = vendor_freq

# # Display the new columns
# print("Region columns added:")
# print(region_dummies.columns.tolist())
# print("\nCategory columns added:")
# print(category_dummies.columns.tolist())
# print("\nVendor frequency encoding sample:")
# print(f_data[['vendor_id', 'vendor_id_freq']].head(10))
# print("\nDataFrame shape:", f_data.shape)

# %%
# Boolean flags are converted to numeric in the feature matrix build above.
pass

# %%
# Flag exact duplicate invoice submissions based on the line-item signature.
duplicate_key_cols = ['vendor_id', 'category', 'region', 'quantity', 'unit_price', 'total_price', 'po_number', 'invoice_number']
f_data['is_duplicate_submission'] = f_data.duplicated(subset=duplicate_key_cols, keep=False)

print("Rows in duplicate submission groups:", f_data['is_duplicate_submission'].sum())
print(
    "Duplicate submission groups:",
    f_data.loc[f_data['is_duplicate_submission']].groupby(duplicate_key_cols).size().reset_index(name='row_count').shape[0]
)

# %%
# Round calculated and stated totals to 2 decimal places, then flag mismatches

expected_total = (f_data['quantity'] * f_data['unit_price']).round(2)
stated_total = f_data['total_price'].round(2)

valid_mask = f_data[['quantity', 'unit_price', 'total_price']].notna().all(axis=1)
disparity = (expected_total - stated_total).abs()

f_data['price_mismatch_flag'] = False
f_data.loc[valid_mask, 'price_mismatch_flag'] = (disparity.loc[valid_mask] > 0.02)

mismatch_rows = f_data.loc[f_data['price_mismatch_flag'], ['quantity', 'unit_price', 'total_price', 'price_mismatch_flag']]

print(f"Rows with mismatched totals: {mismatch_rows.shape[0]}")
mismatch_rows.head()

# %%
key_cols = ['invoice_number', 'po_number', 'timestamp']
amount_cols = ['quantity', 'unit_price', 'total_price', 'total']

amount_variation = f_data.groupby(key_cols)[amount_cols].nunique()
mismatch_groups = amount_variation[(amount_variation > 1).any(axis=1)]

f_data['same_invoice_po_timestamp_diff_amount_flag'] = (
    f_data.set_index(key_cols).index.isin(mismatch_groups.index)
)

print("Rows flagged:", f_data['same_invoice_po_timestamp_diff_amount_flag'].sum())
f_data.loc[
    f_data['same_invoice_po_timestamp_diff_amount_flag'],
    key_cols + amount_cols
].head(10)

# %%
# # Flag same PO repeated with different invoices at the same timestamp
# f_data['same_po_timestamp_diff_invoice_flag'] = (
#     f_data.groupby(['timestamp', 'po_number'])['invoice_number'].transform('nunique') > 1
# )

# print("Rows flagged:", f_data['same_po_timestamp_diff_invoice_flag'].sum())
# # show sample flagged groups
# f_data.loc[f_data['same_po_timestamp_diff_invoice_flag'], ['timestamp', 'po_number', 'invoice_number']].drop_duplicates().head(20)

# %%
# flag rows where computed total differs from stated total_price
f_data['total_vs_total_price_diff'] = (f_data['total'] - f_data['total_price']).abs()
f_data['total_vs_total_price_flag'] = f_data['total_vs_total_price_diff'] > 0.01

print("Rows with total mismatch:", f_data['total_vs_total_price_flag'].sum())
f_data.loc[
    f_data['total_vs_total_price_flag'],
    ['quantity', 'unit_price', 'total', 'total_price', 'total_vs_total_price_diff']
].head()

# %%
# Item purchase counts are not used directly in X, so this step is intentionally left out.
pass


# %%
# Calculate historical mean and std for each vendor-item pair
f_data['avg_quantity_vendor_item'] = (
    f_data.groupby(['vendor_id', 'category'])['quantity']
    .transform('mean')
)

f_data['std_quantity_vendor_item'] = (
    f_data.groupby(['vendor_id', 'category'])['quantity']
    .transform('std')
)


# %%
# Z-score: how far current quantity deviates from historical mean
f_data['std_quantity_vendor_item'] = f_data['std_quantity_vendor_item'].fillna(1).replace(0, 1)
f_data['item_quantity_zscore'] = np.where(
    f_data['quantity'].notna() & f_data['avg_quantity_vendor_item'].notna() & f_data['std_quantity_vendor_item'].notna(),
    (f_data['quantity'] - f_data['avg_quantity_vendor_item']) / f_data['std_quantity_vendor_item'],
    0.0
)


# %%
# The boolean threshold is redundant with item_quantity_zscore and is not used in X.
pass


# %%
print("Item quantity z-score summary:")
print(f_data[['quantity', 'avg_quantity_vendor_item', 'std_quantity_vendor_item', 'item_quantity_zscore']].head())
print("Non-finite item quantity z-scores:", (~np.isfinite(f_data['item_quantity_zscore'])).sum())

# %%
# Vendor's most frequent region (mode)
vendor_region_mode = (
    f_data.groupby('vendor_id')['region']
    .agg(lambda x: x.mode()[0] if not x.mode().empty else None)
    .reset_index(name='vendor_region_mode')
)

# Merge back
f_data = f_data.merge(vendor_region_mode, on='vendor_id', how='left')

# Flag region change
f_data['region_changed'] = f_data['region'] != f_data['vendor_region_mode']


# %%
# Vendor's most common transaction hour
f_data['transaction_hour'] = f_data['timestamp'].dt.hour

vendor_hour_mode = (
    f_data.groupby('vendor_id')['transaction_hour']
    .agg(lambda x: x.mode()[0] if not x.mode().empty else None)
    .reset_index(name='vendor_hour_mode')
)

f_data = f_data.merge(vendor_hour_mode, on='vendor_id', how='left')

# Flag unusual hours
f_data['is_outside_usual_hours'] = f_data['transaction_hour'] != f_data['vendor_hour_mode']


# %%
# Sort by vendor and timestamp
f_data = f_data.sort_values(['vendor_id', 'timestamp'])

# Time difference between consecutive transactions
f_data['inter_transaction_time'] = f_data.groupby('vendor_id')['timestamp'].diff().dt.total_seconds()

# Flag burst activity (e.g., < 300 seconds = 5 minutes)
f_data['is_burst_activity'] = f_data['inter_transaction_time'] < 300


# %%
# Vendor historical mean and std of amounts
f_data['vendor_amount_mean'] = f_data.groupby('vendor_id')['total_price'].transform('mean')
f_data['vendor_amount_std'] = f_data.groupby('vendor_id')['total_price'].transform('std').fillna(1).replace(0, 1)

# Z-score for transaction amount, with a safe fallback so missing input does not create NaN values.
f_data['amount_zscore'] = np.where(
    f_data['total_price'].notna() & f_data['vendor_amount_mean'].notna() & f_data['vendor_amount_std'].notna(),
    (f_data['total_price'] - f_data['vendor_amount_mean']) / f_data['vendor_amount_std'],
    0.0
)

# Flag outliers (e.g., > 3 std deviations)
f_data['is_amount_outlier'] = f_data['amount_zscore'].abs() > 3

# %%
# Step 1: Create a proper date column from year, month, day
f_data['date'] = pd.to_datetime(f_data[['year','month','day']])

# Step 2: Count transactions per vendor per day
daily_counts = (
    f_data.groupby(['vendor_id','date'])
    .size()
    .reset_index(name='daily_txn_count')
)

# Step 3: Compute rolling 7-day average per vendor
daily_counts['rolling_txn_count'] = (
    daily_counts.groupby('vendor_id')['daily_txn_count']
    .transform(lambda x: x.rolling(7, min_periods=1).mean())
)

# Step 4: Merge back into main DataFrame
f_data = f_data.merge(daily_counts, on=['vendor_id','date'], how='left')

# Step 5: Compare rolling count to historical mean
f_data['vendor_txn_mean'] = f_data.groupby('vendor_id')['daily_txn_count'].transform('mean')
f_data['frequency_change_ratio'] = np.where(
    f_data['rolling_txn_count'].notna() & f_data['vendor_txn_mean'].notna() & (f_data['vendor_txn_mean'] != 0),
    f_data['rolling_txn_count'] / f_data['vendor_txn_mean'],
    0.0
)

# Flag spikes/drops
f_data['is_frequency_spike_drop'] = (f_data['frequency_change_ratio'] > 2) | (f_data['frequency_change_ratio'] < 0.5)

# %%
# Step 1: Count transactions per vendor per month
vendor_monthly_activity = (
    f_data.groupby(['vendor_id','year','month'])
    .size()
    .reset_index(name='txn_count_month')
)

# Step 2: Compute vendor’s average monthly activity
vendor_monthly_activity['avg_txn_per_month_vendor'] = (
    vendor_monthly_activity.groupby('vendor_id')['txn_count_month']
    .transform('mean')
)

# Step 3: Merge back into main DataFrame
# Use unique column names to avoid clashes
f_data = f_data.merge(
    vendor_monthly_activity[['vendor_id','year','month','txn_count_month','avg_txn_per_month_vendor']],
    on=['vendor_id','year','month'],
    how='left'
)

# Step 4: Flag deviations from seasonal norms
f_data['deviation_from_avg_vendor_month'] = np.where(
    f_data['txn_count_month'].notna() & f_data['avg_txn_per_month_vendor'].notna(),
    f_data['txn_count_month'] - f_data['avg_txn_per_month_vendor'],
    0.0
)
f_data['flag_seasonal_anomaly_vendor'] = (
    f_data['deviation_from_avg_vendor_month'].abs() > f_data['avg_txn_per_month_vendor'] * 0.5
)

# %%
# Step A: Category quantity z-score per vendor
f_data['category_quantity_mean'] = f_data.groupby(['vendor_id','category'])['quantity'].transform('mean')
f_data['category_quantity_std'] = f_data.groupby(['vendor_id','category'])['quantity'].transform('std').fillna(1).replace(0, 1)

# Avoid division by zero and keep the feature finite even when the input amount fields are missing.
f_data['category_quantity_zscore'] = np.where(
    f_data['quantity'].notna() & f_data['category_quantity_mean'].notna() & f_data['category_quantity_std'].notna(),
    (f_data['quantity'] - f_data['category_quantity_mean']) / f_data['category_quantity_std'],
    0.0
)

# Step B: Excess category purchase flag
f_data['is_excess_category_purchase'] = f_data['category_quantity_zscore'].abs() > 3  # threshold can be tuned

# Step C: New category for vendor flag
# Sort explicitly by vendor and time so first-seen categories reflect chronology rather than incidental row order.
f_data = f_data.sort_values(['vendor_id', 'timestamp'], kind='mergesort')
f_data['is_new_category_for_vendor'] = (
    f_data.groupby('vendor_id')['category'].transform(lambda x: ~x.duplicated())
).astype(int)

# %%
# Add richer categorical and interaction features to improve anomaly coverage.
# Keep the categorical values explicit so the model can learn more from vendor/region/category behavior.
f_data['vendor_id'] = f_data['vendor_id'].fillna('MISSING').astype(str)
f_data['category'] = f_data['category'].fillna('MISSING').astype(str)
f_data['region'] = f_data['region'].fillna('MISSING').astype(str)

vendor_freq = f_data['vendor_id'].value_counts()
category_freq = f_data['category'].value_counts()
region_freq = f_data['region'].value_counts()

f_data['vendor_id_freq'] = f_data['vendor_id'].map(vendor_freq).fillna(0)
f_data['category_freq'] = f_data['category'].map(category_freq).fillna(0)
f_data['region_freq'] = f_data['region'].map(region_freq).fillna(0)

categorical_dummies = pd.get_dummies(
    f_data[['category', 'region']].astype(str),
    prefix=['category', 'region'],
    dummy_na=False
)
f_data = pd.concat([f_data, categorical_dummies], axis=1)

# Create interaction features that capture combined risk signals.
f_data['amount_x_unusual_hours'] = (
    f_data['amount_zscore'] * f_data['is_outside_usual_hours'].astype(float)
)
f_data['amount_x_new_category'] = (
    f_data['amount_zscore'] * f_data['is_new_category_for_vendor'].astype(float)
)
f_data['duplicate_x_price_mismatch'] = (
    f_data['is_duplicate_submission'].astype(float) * f_data['price_mismatch_flag'].astype(float)
)
f_data['duplicate_x_amount_outlier'] = (
    f_data['is_duplicate_submission'].astype(float) * f_data['is_amount_outlier'].astype(float)
)

# Add PO/invoice pattern flags for repeated submission behavior.
po_invoice_group_counts = (
    f_data.groupby(['po_number', 'invoice_number'])
    .size()
    .reset_index(name='po_invoice_group_count')
)
f_data = f_data.merge(po_invoice_group_counts, on=['po_number', 'invoice_number'], how='left')

f_data['repeat_po_invoice_group'] = f_data['po_invoice_group_count'] > 1
f_data['same_po_multiple_invoices'] = (
    f_data.groupby('po_number')['invoice_number'].transform('nunique') > 1
)

print("Added categorical encoding and interaction features.")

# %%
# Step 1: Collect all engineered features into a single matrix
feature_cols = [
    # Missing-data signal
    'has_missing_amount_fields',
    'is_duplicate_submission',

    # Invoice integrity
    'price_mismatch_flag',
    'same_invoice_po_timestamp_diff_amount_flag',
    'total_vs_total_price_flag',

    # Item/Category
    'item_quantity_zscore',
    'category_quantity_zscore',
    'is_new_category_for_vendor',

    # Vendor Geography
    'region_changed',

    # Vendor Temporal
    'is_outside_usual_hours',
    'is_burst_activity',

    # Vendor Amount
    'amount_zscore',
    'is_amount_outlier',

    # Vendor Frequency & Seasonality
    'frequency_change_ratio',
    'deviation_from_avg_vendor_month',

    # Categorical and interaction features
    'vendor_id_freq',
    'category_freq',
    'region_freq',
    'amount_x_unusual_hours',
    'amount_x_new_category',
    'duplicate_x_price_mismatch',
    'duplicate_x_amount_outlier',
    'repeat_po_invoice_group',
    'same_po_multiple_invoices',
]

# Add one-hot encoded category/region columns to the feature matrix.
feature_cols.extend([col for col in f_data.columns if col.startswith(('category_', 'region_'))])

# Remove any duplicate columns that may have appeared after rerunning earlier cells.
f_data = f_data.loc[:, ~f_data.columns.duplicated()].copy()
feature_cols = [col for col in dict.fromkeys(feature_cols) if col in f_data.columns]

# Step 2: Create the feature matrix (X)
X = f_data.reindex(columns=feature_cols).copy()

# Step 3: Ensure all boolean flags are numeric (0/1) and keep the matrix finite.
X = X.astype(float)
X = X.replace([np.inf, -np.inf], np.nan)

print("Final feature columns:", len(feature_cols))
print(X.head())

# %%
print("X missing values:\n", X.isna().sum())
print("X infinite values:\n", np.isinf(X).sum())
print("X summary:\n")
print(X.describe())

if X.isna().any().any():
    raise ValueError("X still contains NaN values before IsolationForest.")
if np.isinf(X).values.any():
    raise ValueError("X still contains infinite values before IsolationForest.")

# %%
from sklearn.ensemble import IsolationForest

# Stable Isolation Forest entry point. Do not edit this cell.
def run_isolation_forest(data_frame=None):
    global f_data, X, iso

    if data_frame is None:
        if 'f_data' not in globals():
            data_frame = pd.read_excel(r"C:\Users\royr\Downloads\Sample_fraud_analytics_intern_v1.1.xlsx")
        else:
            data_frame = f_data.copy()
    else:
        data_frame = data_frame.copy()

    feature_cols = [
        'has_missing_amount_fields',
        'is_duplicate_submission',
        'price_mismatch_flag',
        'same_invoice_po_timestamp_diff_amount_flag',
        'total_vs_total_price_flag',
        'item_quantity_zscore',
        'category_quantity_zscore',
        'is_new_category_for_vendor',
        'region_changed',
        'is_outside_usual_hours',
        'is_burst_activity',
        'amount_zscore',
        'is_amount_outlier',
        'frequency_change_ratio',
        'deviation_from_avg_vendor_month',
        'vendor_id_freq',
        'category_freq',
        'region_freq',
        'amount_x_unusual_hours',
        'amount_x_new_category',
        'duplicate_x_price_mismatch',
        'duplicate_x_amount_outlier',
        'repeat_po_invoice_group',
        'same_po_multiple_invoices',
    ]
    feature_cols = [col for col in feature_cols if col in data_frame.columns]
    feature_cols.extend([col for col in data_frame.columns if col.startswith(('category_', 'region_'))])
    feature_cols = [col for col in dict.fromkeys(feature_cols) if col in data_frame.columns]

    data_frame = data_frame.loc[:, ~data_frame.columns.duplicated()].copy()
    X = data_frame.reindex(columns=feature_cols).copy()
    X = X.astype(float).replace([np.inf, -np.inf], np.nan).fillna(0.0)
    X = X.loc[:, ~X.columns.duplicated()].copy()

    iso = IsolationForest(
        n_estimators=200,
        contamination="auto",
        random_state=42,
        n_jobs=-1
    )
    iso.fit(X)

    data_frame['anomaly_score'] = -iso.score_samples(X)
    data_frame['is_anomaly'] = iso.predict(X) == -1

    f_data = data_frame
    return f_data, X, iso


f_data, X, iso = run_isolation_forest()

print("Stable IsolationForest configuration: contamination=0.07, n_estimators=200, random_state=42")
print("Anomaly rate:", f_data['is_anomaly'].mean())
print(f_data[['is_anomaly', 'anomaly_score']].head())


# %%
# Inspect the flagged anomalies
anomalies = f_data.loc[f_data['is_anomaly']].copy()
print("Number of flagged anomalies:", len(anomalies))

if len(anomalies) > 0:
    display_cols = [
        'vendor_id', 'category', 'region', 'quantity', 'unit_price', 'total_price',
        'invoice_number', 'po_number', 'timestamp', 'price_mismatch_flag',
        'same_invoice_po_timestamp_diff_amount_flag', 'total_vs_total_price_flag',
        'has_missing_amount_fields', 'is_duplicate_submission',
        'item_quantity_zscore', 'category_quantity_zscore', 'amount_zscore',
        'frequency_change_ratio', 'deviation_from_avg_vendor_month', 'anomaly_score'
    ]
    print(anomalies[display_cols].head(20))
else:
    print("No anomalies were flagged.")

# %%
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Build a review-ready Excel export with row-level anomaly type coloring.
anomalies = f_data.loc[f_data['is_anomaly']].copy()

if len(anomalies) == 0:
    print("No anomalies were flagged, so no Excel export was created.")
else:
    anomalies['suspicion_score'] = (
        anomalies['anomaly_score']
        + 0.5 * anomalies['is_duplicate_submission'].astype(float)
        + 0.5 * anomalies['price_mismatch_flag'].astype(float)
        + 0.5 * anomalies['same_invoice_po_timestamp_diff_amount_flag'].astype(float)
        + 0.5 * anomalies['total_vs_total_price_flag'].astype(float)
        + 0.5 * anomalies['has_missing_amount_fields'].astype(float)
        + 0.15 * np.abs(anomalies['item_quantity_zscore'])
        + 0.15 * np.abs(anomalies['category_quantity_zscore'])
        + 0.15 * np.abs(anomalies['amount_zscore'])
    )

    def classify_anomaly(row: pd.Series) -> str:
        has_missing = bool(row.get('has_missing_amount_fields', False))
        duplicate = bool(row.get('is_duplicate_submission', False))
        integrity = bool(row.get('price_mismatch_flag', False)) or bool(row.get('total_vs_total_price_flag', False)) or bool(row.get('same_invoice_po_timestamp_diff_amount_flag', False))
        outlier = (
            abs(float(row.get('item_quantity_zscore', 0))) > 3
            or abs(float(row.get('category_quantity_zscore', 0))) > 3
            or abs(float(row.get('amount_zscore', 0))) > 3
        )

        if has_missing and duplicate and integrity:
            return 'Missing + Duplicate + Integrity'
        if has_missing:
            return 'Missing Amount'
        if duplicate and integrity:
            return 'Duplicate + Integrity'
        if duplicate:
            return 'Duplicate'
        if integrity:
            return 'Integrity Mismatch'
        if outlier:
            return 'Statistical Outlier'
        return 'Behavioral'

    anomalies['anomaly_type'] = anomalies.apply(classify_anomaly, axis=1)

    export_cols = [
        'vendor_id', 'category', 'region', 'quantity', 'unit_price', 'total_price',
        'invoice_number', 'po_number', 'timestamp', 'price_mismatch_flag',
        'same_invoice_po_timestamp_diff_amount_flag', 'total_vs_total_price_flag',
        'has_missing_amount_fields', 'is_duplicate_submission',
        'item_quantity_zscore', 'category_quantity_zscore', 'amount_zscore',
        'suspicion_score', 'anomaly_type'
    ]
    export_df = anomalies[export_cols].copy()

    output_path = r"C:\Users\royr\Downloads\anomaly_report_colored.xlsx"
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='anomalies', index=False)

    wb = load_workbook(output_path)
    ws = wb['anomalies']

    color_map = {
        'Missing Amount': 'D5F5E3',
        'Duplicate': 'FADBD8',
        'Integrity Mismatch': 'FDEBD0',
        'Statistical Outlier': 'D6EAF8',
        'Duplicate + Integrity': 'F5B7B1',
        'Missing + Duplicate + Integrity': 'E8DAEF',
        'Behavioral': 'F4F6F6',
    }

    header_fill = PatternFill(fill_type='solid', fgColor='D5DBDB')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx in range(2, ws.max_row + 1):
        anomaly_type = ws.cell(row=row_idx, column=export_df.columns.get_loc('anomaly_type') + 1).value or 'Behavioral'
        fill = PatternFill(fill_type='solid', fgColor=color_map.get(anomaly_type, 'F4F6F6'))
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Excel exported to: {output_path}")
    print(f"Rows written: {len(export_df)}")


# %%
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
from sklearn.preprocessing import StandardScaler

# Use a copy of X so the original feature matrix remains unchanged.
X_ae_work = X.copy().astype(float)
scaler_ae = StandardScaler()
X_ae_scaled = scaler_ae.fit_transform(X_ae_work)

class SimpleAutoencoder(nn.Module):
    def __init__(self, input_dim):
        super().__init__()
        self.encoder = nn.Sequential(
            nn.Linear(input_dim, 32),
            nn.ReLU(),
            nn.Linear(32, 16),
            nn.ReLU(),
        )
        self.decoder = nn.Sequential(
            nn.Linear(16, 32),
            nn.ReLU(),
            nn.Linear(32, input_dim),
        )

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z)


X_tensor = torch.tensor(X_ae_scaled, dtype=torch.float32)
dataset = TensorDataset(X_tensor)
loader = DataLoader(dataset, batch_size=64, shuffle=True)

model = SimpleAutoencoder(X_ae_scaled.shape[1])
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=1e-3)

for epoch in range(20):
    model.train()
    for batch, in loader:
        optimizer.zero_grad()
        recon = model(batch)
        loss = criterion(recon, batch)
        loss.backward()
        optimizer.step()

model.eval()
with torch.no_grad():
    recon = model(X_tensor)
    recon_error = ((X_tensor - recon) ** 2).mean(dim=1).numpy()

threshold = np.percentile(recon_error, 93)
f_data['ae_anomaly'] = (recon_error > threshold).astype(bool)

comparison_df = pd.DataFrame({
    'iso': f_data['is_anomaly'].astype(bool),
    'autoencoder': f_data['ae_anomaly'].astype(bool),
})

print("Autoencoder anomaly rate:", comparison_df['autoencoder'].mean())
print("Autoencoder flagged:", int(comparison_df['autoencoder'].sum()))
print("Overlap with ISO:", int((comparison_df['iso'] & comparison_df['autoencoder']).sum()))
print(comparison_df.sum())


# %%
from pyod.models.hbos import HBOS
from sklearn.preprocessing import StandardScaler

# Use a copy of X so the original feature matrix remains unchanged.
X_hbos_work = X.copy().astype(float)
scaler_hbos = StandardScaler()
X_hbos_work = scaler_hbos.fit_transform(X_hbos_work)

hbos = HBOS(contamination=0.07, n_bins=20)
hbos.fit(X_hbos_work)

scores = np.nan_to_num(hbos.decision_scores_, nan=0.0)
threshold = np.percentile(scores, 93)
f_data['hbos_anomaly'] = (scores > threshold).astype(bool)

comparison_df = pd.DataFrame({
    'iso': f_data['is_anomaly'].astype(bool),
    'hbos': f_data['hbos_anomaly'].astype(bool),
})

print("HBOS anomaly rate:", comparison_df['hbos'].mean())
print("HBOS flagged:", int(comparison_df['hbos'].sum()))
print("Overlap with ISO:", int((comparison_df['iso'] & comparison_df['hbos']).sum()))
print("Score summary:")
print(pd.Series(scores).describe())
print(comparison_df.sum())


# %%
from pyod.models.copod import COPOD

# Use a copy of X so the original feature matrix remains unchanged.
X_copod_work = X.copy().astype(float)

copod = COPOD(contamination=0.07)
copod.fit(X_copod_work)

scores = np.nan_to_num(copod.decision_scores_, nan=0.0)
threshold = np.percentile(scores, 93)
f_data['copod_anomaly'] = (scores > threshold).astype(bool)

comparison_df = pd.DataFrame({
    'iso': f_data['is_anomaly'].astype(bool),
    'copod': f_data['copod_anomaly'].astype(bool),
})

print("COPOD anomaly rate:", comparison_df['copod'].mean())
print("COPOD flagged:", int(comparison_df['copod'].sum()))
print("Overlap with ISO:", int((comparison_df['iso'] & comparison_df['copod']).sum()))
print("Score summary:")
print(pd.Series(scores).describe())
print(comparison_df.sum())


# %%
from sklearn.neighbors import LocalOutlierFactor

# Use a copy of X so the original feature matrix remains unchanged.
X_lof_work = X.copy().astype(float)

lof = LocalOutlierFactor(n_neighbors=20, contamination=0.07)
lof_predictions = lof.fit_predict(X_lof_work)
f_data['lof_anomaly'] = (lof_predictions == -1).astype(bool)

comparison_df = pd.DataFrame({
    'iso': f_data['is_anomaly'].astype(bool),
    'autoencoder': f_data.get('ae_anomaly', pd.Series([False] * len(f_data))).astype(bool),
    'hbos': f_data.get('hbos_anomaly', pd.Series([False] * len(f_data))).astype(bool),
    'copod': f_data.get('copod_anomaly', pd.Series([False] * len(f_data))).astype(bool),
    'lof': f_data['lof_anomaly'].astype(bool),
})

print("LOF anomaly rate:", comparison_df['lof'].mean())
print("LOF flagged:", int(comparison_df['lof'].sum()))
print("Overlap with ISO:", int((comparison_df['iso'] & comparison_df['lof']).sum()))
print(comparison_df.sum())

from matplotlib_venn import venn2

# Create Venn diagrams for each comparison with Isolation Forest.
fig, axes = plt.subplots(1, 4, figsize=(24, 5))

venn2([set(comparison_df.loc[comparison_df['iso']].index), set(comparison_df.loc[comparison_df['autoencoder']].index)],
      set_labels=('Isolation Forest', 'Autoencoder'), ax=axes[0])
axes[0].set_title('ISO vs Autoencoder')

venn2([set(comparison_df.loc[comparison_df['iso']].index), set(comparison_df.loc[comparison_df['hbos']].index)],
      set_labels=('Isolation Forest', 'HBOS'), ax=axes[1])
axes[1].set_title('ISO vs HBOS')

venn2([set(comparison_df.loc[comparison_df['iso']].index), set(comparison_df.loc[comparison_df['copod']].index)],
      set_labels=('Isolation Forest', 'COPOD'), ax=axes[2])
axes[2].set_title('ISO vs COPOD')

venn2([set(comparison_df.loc[comparison_df['iso']].index), set(comparison_df.loc[comparison_df['lof']].index)],
      set_labels=('Isolation Forest', 'LOF'), ax=axes[3])
axes[3].set_title('ISO vs LOF')

plt.tight_layout()
plt.show()

print("Overlap counts:")
print({
    'ISO vs Autoencoder': int((comparison_df['iso'] & comparison_df['autoencoder']).sum()),
    'ISO vs HBOS': int((comparison_df['iso'] & comparison_df['hbos']).sum()),
    'ISO vs COPOD': int((comparison_df['iso'] & comparison_df['copod']).sum()),
    'ISO vs LOF': int((comparison_df['iso'] & comparison_df['lof']).sum()),
})



